"""
This codes used to parse correspondence files uploaded by users. 
They are served to read the correspondence data and validate its format:
- Each line should have 2 or 3 columns separated by tabs, commas, semicolons, or multiple spaces.
- The first column is the identifier, the second is the display name, and the optional third.
- ALAWAYS consider the first row as a HEADER and skip it for .xlsx files.
"""


import re
from typing import List, Tuple
from openpyxl import load_workbook


# Parses .txt/.csv correspondence files
def parse_correspondence_text(raw_text: str):
    """    
    :param raw_text: Description
    :type raw_text: str
    """
    rows = []
    errors = []
    seen = set()

    for lineno, line in enumerate(raw_text.splitlines(), start=1):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        parts = re.split(r"[,\t;]+|\s{2,}", line)
        parts = [p.strip() for p in parts if p.strip()] 

        if len(parts) not in (2, 3):
            errors.append(f"Line {lineno}: expected 2 or 3 columns, got {len(parts)} -> {line}")
            continue

        identifier, display_name = parts[0], parts[1]
        entry_type = parts[2] if len(parts) == 3 else ""

        if identifier in seen:
            errors.append(f"Line {lineno}: duplicate identifier '{identifier}' in file.")
            continue
        seen.add(identifier)

        rows.append((identifier, display_name, entry_type))
    
    return rows, errors


# Parses .xlsx correspondence files
def parse_correspondence_xlsx(
    uploaded_file,
) -> tuple[list[tuple[str, str, str]], list[str]]:
    rows: list[tuple[str, str, str]] = []
    errors: list[str] = []
    seen = set()

    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active  # always use first sheet

    for lineno, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # === Always skip first row (header) ===
        if lineno == 1:
            continue

        # skip empty rows
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        identifier = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        display_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        entry_type = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not identifier or not display_name:
            errors.append(f"Line {lineno}: identifier and display_name are required.")
            continue

        if identifier in seen:
            errors.append(f"Line {lineno}: duplicate identifier '{identifier}' in file.")
            continue
        seen.add(identifier)

        rows.append((identifier, display_name, entry_type))

    return rows, errors