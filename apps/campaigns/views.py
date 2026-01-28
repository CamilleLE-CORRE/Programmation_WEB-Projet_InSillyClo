import csv
import io
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CampaignTemplate
from .forms import CampaignTemplateForm  # You'll need to create this form


def build_excel_response(template, filename):
    """Attempt to build a styled XLSX response for the given template.
    Returns an HttpResponse with XLSX content, or None if openpyxl is not available."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return None

    # Build rows depending on template type (same structure as CSV version)
    rows = []
    rows.append(["Assembly settings"])                                           # A1
    rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
    rows.append(["Name", template.name or ""])                                 # A3, B3
    rows.append(["Output separator", template.separator or ""])               # A4, B4
    # rows 5-8 empty
    rows.extend([[] for _ in range(4)])

    # Determine headers and typed-specific content
    if template.template_type == 'simple':
        # Row 9: Assembly composition and headers C..J
        header = ["Assembly composition", "Part name ->"] + [f"input Plasmid {i}" for i in range(1, 9)]
        rows.append(header)
        # Row 10,11,12,13,14 as simple layout
        row10 = ["", "Part types ->"] + ["" for _ in range(8)]
        rows.append(row10)
        row11 = ["", "Is optional part ->"] + ["True" for _ in range(8)]
        rows.append(row11)
        row12 = ["", "Part name should be in output name ->"] + ["True" for _ in range(8)]
        rows.append(row12)
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)
    else:  # typed
        header = ["Assembly composition", "Part name ->", "ConL", "Promoter", "Gene", "Terminator", "ConR", "Backbone", "", ""]
        rows.append(header)
        row10 = ["", "Part types ->", "1", "2,[2a, 2b]", "3,[3a, 3b]", "4,[4a, 4b]", "5", "678,[6, 7, 8]", "", ""]
        rows.append(row10)
        row11 = ["", "Is optional part ->", "True", "False", "False", "False", "True", "False", "", ""]
        rows.append(row11)
        row12 = ["", "Part name should be in output name ->", "False", "True", "True", "True", "False", "False", "", ""]
        rows.append(row12)
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)

    # Create workbook and write rows
    wb = Workbook()
    ws = wb.active
    max_cols = 10  # columns A-J

    # Define styles
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # light green
    blue_fill = PatternFill(start_color="CCEEFF", end_color="CCEEFF", fill_type="solid")   # light blue
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Determine green area ranges
    if template.template_type == 'simple':
        green_cols = set(range(3, 11))  # C..J
    else:
        green_cols = set(range(3, 9))   # C..H
    green_rows = set(range(9, 14))      # 9..13 inclusive

    for r_idx, row in enumerate(rows, start=1):
        for c_idx in range(1, max_cols + 1):
            val = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Apply green fill when in the specified ranges
            if r_idx in green_rows and c_idx in green_cols:
                cell.fill = green_fill
            elif val not in (None, ""):
                # light blue for other cells with content
                cell.fill = blue_fill

    # Vertical border between columns B(2) and C(3): set left border on column C
    for r in range(1, len(rows) + 1):
        cell = ws.cell(row=r, column=3)
        # preserve existing borders if any (we overwrite left)
        cell.border = Border(left=medium, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)

    # Horizontal border between row 14 and 15: set bottom border on row 14
    for c in range(1, max_cols + 1):
        cell = ws.cell(row=14, column=c)
        cell.border = Border(bottom=medium, left=cell.border.left, right=cell.border.right, top=cell.border.top)

    # Draw vertical borders only (no horizontal borders) for rows 15..50.
    # We apply left+right thin borders to create vertical separators between columns,
    # but omit top/bottom so there are no horizontal lines between rows.
    for r in range(15, 51):
        for c in range(1, max_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin)

    # Set reasonable column widths
    for c in range(1, max_cols + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 18

    # Save workbook to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


# View for the main templates page
def template_list(request):
    # Get all public templates (always visible)
    public_templates = CampaignTemplate.objects.filter(is_public=True)
    
    # Get user's templates if authenticated
    user_templates = None
    if request.user.is_authenticated:
        user_templates = CampaignTemplate.objects.filter(owner=request.user)
    
    context = {
        'public_templates': public_templates,
        'user_templates': user_templates,
    }
    return render(request, 'campaigns/template_list.html', context)


# View to download a predefined (public) template
def download_public_template(request, template_id):
    # Get the public template or return 404
    template = get_object_or_404(CampaignTemplate, id=template_id, is_public=True)

    fmt = (request.GET.get('format') or '').lower()

    rows = []
    rows.append(["Assembly settings"])                                           # A1
    rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
    rows.append(["Name", template.name or ""])                                 # A3, B3
    rows.append(["Output separator", template.separator or ""])               # A4, B4
    rows.extend([[] for _ in range(4)])
    header = ["Assembly composition", "Part name ->"] + [
        "ConL", "Promoter", "Gene", "Terminator", "ConR", "Backbone", "", ""
    ]
    rows.append(header)
    row10 = ["", "Part types ->"] + [
        "1", "2,[2a, 2b]", "3,[3a, 3b]", "4,[4a, 4b]", "5", "678,[6, 7, 8]", "", ""
    ]
    rows.append(row10)
    row11 = ["", "Is optional part ->"] + ["True", "False", "False", "False", "True", "False", "", ""]
    rows.append(row11)
    row12 = ["", "Part name should be in output name ->"] + ["False", "True", "True", "True", "False", "False", "", ""]
    rows.append(row12)
    row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
    rows.append(row13)
    row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
    rows.append(row14)
    
    for r in rows:
        writer.writerow(r)

    return response

    # Prefer building an XLSX with formatting when available
    xlsx_resp = build_excel_response(template, f"template_{template.name}")
    if xlsx_resp is not None:
        return xlsx_resp
    # If the caller explicitly requested XLSX, but we couldn't build it (e.g. openpyxl missing),
    # respond with an explicit error so the user knows why a CSV was not provided.
    if fmt == 'xlsx':
        return HttpResponse(
            "XLSX export is not available on the server (openpyxl is not installed). Please use ?format=csv to download a CSV instead.",
            status=501,
            content_type='text/plain'
        )

    # Fallback to CSV if openpyxl not available
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="template_{template.name}.csv"'
    writer = csv.writer(response)    
    # Special CSV layout for simple templates (spreadsheet style)
    if template.template_type == 'simple':
        # Build 14 rows with up to 10 columns (A-J)
        rows = []
        rows.append(["Assembly settings"])                                           # A1
        rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
        rows.append(["Name", template.name or ""])                                 # A3, B3
        rows.append(["Output separator", template.separator or ""])               # A4, B4
        # rows 5-8 empty
        rows.extend([[] for _ in range(4)])
        # Row 9: Assembly composition and headers
        header = ["Assembly composition", "Part name ->"]
        # add input plasmid headers from C to J (1..8)
        for i in range(1, 9):
            header.append(f"input Plasmid {i}")
        rows.append(header)
        # Row 10: Part types label
        row10 = ["", "Part types ->"] + ["" for _ in range(8)]
        rows.append(row10)
        # Row 11: Is optional part -> with True in C..J
        row11 = ["", "Is optional part ->"] + ["True" for _ in range(8)]
        rows.append(row11)
        # Row 12: Part name should be in output name -> with True in C..J
        row12 = ["", "Part name should be in output name ->"] + ["True" for _ in range(8)]
        rows.append(row12)
        # Row 13: Part separator ->
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        # Row 14: Output plasmid id and OutputType (optional) and down arrows in C..J
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)

        # Write rows to CSV
        for r in rows:
            writer.writerow(r)
    else:  # typed
        rows = []
        rows.append(["Assembly settings"])                                           # A1
        rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
        rows.append(["Name", template.name or ""])                                 # A3, B3
        rows.append(["Output separator", template.separator or ""])               # A4, B4
        # rows 5-8 empty
        rows.extend([[] for _ in range(4)])
        # Row 9: Assembly composition and typed headers in C..H
        header = ["Assembly composition", "Part name ->"] + [
            "ConL", "Promoter", "Gene", "Terminator", "ConR", "Backbone", "", ""
        ]
        rows.append(header)
        # Row 10: numbers / part types info in C..H
        row10 = ["", "Part types ->"] + [
            "1", "2,[2a, 2b]", "3,[3a, 3b]", "4,[4a, 4b]", "5", "678,[6, 7, 8]", "", ""
        ]
        rows.append(row10)
        # Row 11: Is optional part -> with True/False in C..H
        row11 = ["", "Is optional part ->"] + ["True", "False", "False", "False", "True", "False", "", ""]
        rows.append(row11)
        # Row 12: Part name should be in output name -> with True/False in C..H
        row12 = ["", "Part name should be in output name ->"] + ["False", "True", "True", "True", "False", "False", "", ""]
        rows.append(row12)
        # Row 13: Part separator -> (keep blanks)
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        # Row 14: Output plasmid id and OutputType (optional) and down arrows in C..J
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)
        for r in rows:
            writer.writerow(r)
    
    return response


# View to create a new template (authenticated users only)
@login_required
def create_template(request):
    if request.method == 'POST':
        form = CampaignTemplateForm(request.POST)
        if form.is_valid():
            # Create and save the template, then redirect to the list with a success message
            template = form.save(commit=False)
            template.owner = request.user
            template.save()
            messages.success(request, f'The template "{template.name}" has been created !')
            return redirect('campaigns:template_list')
        else:
            messages.error(request, 'Correct the errors below')
    else:
        form = CampaignTemplateForm()

    context = {
        'form': form,
    }
    return render(request, 'campaigns/create_template.html', context)


# View to download a user's created template
@login_required
def download_template(request, template_id):
    # Get the template (must be owned by user or public)
    template = get_object_or_404(
        CampaignTemplate, 
        id=template_id
    )
    
    # Check permissions: user must own the template or it must be public
    if template.owner != request.user and not template.is_public:
        messages.error(request, 'You do not have permission to download this template.')
        return redirect('campaigns:template_list')
    
    fmt = (request.GET.get('format') or '').lower()
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="template_{template.name}.csv"'
        writer = csv.writer(response)
        # Special CSV layout for simple templates (spreadsheet style)
        if template.template_type == 'simple':
            rows = []
            rows.append(["Assembly settings"])                                           # A1
            rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
            rows.append(["Name", template.name or ""])                                 # A3, B3
            rows.append(["Output separator", template.separator or ""])               # A4, B4
            # rows 5-8 empty
            rows.extend([[] for _ in range(4)])
            # Row 9: Assembly composition and headers
            header = ["Assembly composition", "Part name ->"]
            for i in range(1, 9):
                header.append(f"input Plasmid {i}")
            rows.append(header)
            # Row 10: Part types label
            row10 = ["", "Part types ->"] + ["" for _ in range(8)]
            rows.append(row10)
            # Row 11: Is optional part -> with True in C..J
            row11 = ["", "Is optional part ->"] + ["True" for _ in range(8)]
            rows.append(row11)
            # Row 12: Part name should be in output name -> with True in C..J
            row12 = ["", "Part name should be in output name ->"] + ["True" for _ in range(8)]
            rows.append(row12)
            # Row 13: Part separator ->
            row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
            rows.append(row13)
            # Row 14: Output plasmid id and OutputType (optional) and down arrows in C..J
            row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
            rows.append(row14)

            # Write rows to CSV
            for r in rows:
                writer.writerow(r)
        else:
            rows = []
            rows.append(["Assembly settings"])                                           # A1
            rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
            rows.append(["Name", template.name or ""])                                 # A3, B3
            rows.append(["Output separator", template.separator or ""])               # A4, B4
            # rows 5-8 empty
            rows.extend([[] for _ in range(4)])
            # Row 9: Assembly composition and typed headers in C..H
            header = ["Assembly composition", "Part name ->"] + [
                "ConL", "Promoter", "Gene", "Terminator", "ConR", "Backbone", "", ""
            ]
            rows.append(header)
            # Row 10: numbers / part types info in C..H
            row10 = ["", "Part types ->"] + [
                "1", "2,[2a, 2b]", "3,[3a, 3b]", "4,[4a, 4b]", "5", "678,[6, 7, 8]", "", ""
            ]
            rows.append(row10)
            # Row 11: Is optional part -> with True/False in C..H
            row11 = ["", "Is optional part ->"] + ["True", "False", "False", "False", "True", "False", "", ""]
            rows.append(row11)
            # Row 12: Part name should be in output name -> with True/False in C..H
            row12 = ["", "Part name should be in output name ->"] + ["False", "True", "True", "True", "False", "False", "", ""]
            rows.append(row12)
            # Row 13: Part separator -> (keep blanks)
            row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
            rows.append(row13)
            # Row 14: Output plasmid id and OutputType (optional) and down arrows in C..J
            row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
            rows.append(row14)
            for r in rows:
                writer.writerow(r)
        return response



    # Prefer building an XLSX with formatting when available
    xlsx_resp = build_excel_response(template, f"template_{template.name}")
    if xlsx_resp is not None:
        return xlsx_resp
    # If the caller explicitly requested XLSX, but we couldn't build it (e.g. openpyxl missing),
    # respond with an explicit error so the user knows why a CSV was not provided.
    if fmt == 'xlsx':
        return HttpResponse(
            "XLSX export is not available on the server (openpyxl is not installed). Please use ?format=csv to download a CSV instead.",
            status=501,
            content_type='text/plain'
        )

    # Fallback to CSV if openpyxl not available
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="template_{template.name}.csv"'
    
    writer = csv.writer(response)
    
    # Special CSV layout for simple templates (spreadsheet style)
    if template.template_type == 'simple':
        rows = []
        rows.append(["Assembly settings"])                                           # A1
        rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
        rows.append(["Name", template.name or ""])                                 # A3, B3
        rows.append(["Output separator", template.separator or ""])               # A4, B4
        # rows 5-8 empty
        rows.extend([[] for _ in range(4)])
        # Row 9: Assembly composition and headers
        header = ["Assembly composition", "Part name ->"]
        for i in range(1, 9):
            header.append(f"input Plasmid {i}")
        rows.append(header)
        # Row 10: Part types label
        row10 = ["", "Part types ->"] + ["" for _ in range(8)]
        rows.append(row10)
        # Row 11: Is optional part -> with True in C..J
        row11 = ["", "Is optional part ->"] + ["True" for _ in range(8)]
        rows.append(row11)
        # Row 12: Part name should be in output name -> with True in C..J
        row12 = ["", "Part name should be in output name ->"] + ["True" for _ in range(8)]
        rows.append(row12)
        # Row 13: Part separator ->
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        # Row 14: Output plasmid id and OutputType (optional) and down arrows in C..J
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)

        # Write rows to CSV
        for r in rows:
            writer.writerow(r)
    else:
        writer = csv.writer(response)
        rows = []
        rows.append(["Assembly settings"])                                           # A1
        rows.append(["Restriction enzyme", template.restriction_enzyme or ""])     # A2, B2
        rows.append(["Name", template.name or ""])                                 # A3, B3
        rows.append(["Output separator", template.separator or ""])               # A4, B4
        # rows 5-8 empty
        rows.extend([[] for _ in range(4)])
        header = ["Assembly composition", "Part name ->"] + [
            "ConL", "Promoter", "Gene", "Terminator", "ConR", "Backbone", "", ""
        ]
        rows.append(header)
        row10 = ["", "Part types ->"] + [
            "1", "2,[2a, 2b]", "3,[3a, 3b]", "4,[4a, 4b]", "5", "678,[6, 7, 8]", "", ""
        ]
        rows.append(row10)
        row11 = ["", "Is optional part ->"] + ["True", "False", "False", "False", "True", "False", "", ""]
        rows.append(row11)
        row12 = ["", "Part name should be in output name ->"] + ["False", "True", "True", "True", "False", "False", "", ""]
        rows.append(row12)
        row13 = ["", "Part separator ->"] + ["" for _ in range(8)]
        rows.append(row13)
        row14 = ["Output plasmid id ↓", "OutputType (optional) ↓"] + ["↓" for _ in range(8)]
        rows.append(row14)
        for r in rows:
            writer.writerow(r)

    return response


@login_required
def delete_template(request, template_id):
    """Supprime un template appartenant à l'utilisateur connecté"""
    
    # Verify that the template does exist and is a private template
    template = get_object_or_404(
        CampaignTemplate, 
        id=template_id, 
        owner=request.user
    )
    
    if request.method == 'POST':
        template_name = template.name
        template.delete()
        messages.success(request, f"The template '{template_name}' has been deleted with sucess.")
    
    return redirect('campaigns:template_list')

@login_required
def edit_template(request, template_id):
    template = get_object_or_404(
        CampaignTemplate,
        id=template_id,
        owner=request.user
    )

    if request.method == 'POST':
        form = CampaignTemplateForm(request.POST, instance=template)
        if form.is_valid():
            form.save()
            messages.success(request, f'Template "{template.name}" updated.')
            return redirect('campaigns:template_list')
        else:
            messages.error(request, 'Correct the errors below.')
    else:
        form = CampaignTemplateForm(instance=template)

    return render(request, 'campaigns/edit_template.html', {
        'form': form,
        'template': template,
    })

@login_required
def replace_template_xlsx(request, template_id):
    template = get_object_or_404(
        CampaignTemplate,
        id=template_id,
        owner=request.user
    )

    if request.method == 'POST' and request.FILES.get('xlsx'):
        # TODO: parser le fichier si besoin
        template.updated_at = timezone.now()
        template.save()
        messages.success(request, 'Template replaced using new XLSX file.')
        return redirect('campaigns:template_list')

    return render(request, 'campaigns/replace_template_xlsx.html', {
        'template': template,
    })
